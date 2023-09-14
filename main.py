import tkinter as tk
import customtkinter
import openpyxl
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.edge.options import Options
from time import sleep
from PIL import ImageTk, Image


def scrape_website():
    #Variaveis de Pesquisa
    numero_oab = numero_oab_entry.get()
    sigla_estado = sigla_estado_entry.get()

    try:

        pegar_desktop = os.path.expanduser("~\Desktop")
        caminho_do_excel_raiz = os.path.join(pegar_desktop,"Processos.xlsx")


        def print_step(message):
            print(f"{message}")

        #Navegador em modo Headless(Sem abrir a interface gráfica)
        edge_options = Options()
        #edge_options.add_argument('--headless')
        edge_options.add_argument('window-size=1420x1080')

        # Guardar no excel e separar por processos.
        print_step(f"Criando o arquivo excel na Area de Trabalho...")
        workbook_criar = openpyxl.Workbook()
        workbook_criar.save(caminho_do_excel_raiz)

        # entrar no site da - https://pje1g.trf1.jus.br/consultapublica/ConsultaPublica/listView.seam.
        print_step("Abrindo o site...")
        driver = webdriver.Edge(options=edge_options)
        driver.get('https://pje1g.trf1.jus.br/consultapublica/ConsultaPublica/listView.seam')
        sleep(5)

        # Digitar número da OAB
        print_step("Preenchendo numero da OAB...") 
        campo_oab = driver.find_element(By.XPATH,"//input[@id='fPP:Decoration:numeroOAB']")
        campo_oab.send_keys(numero_oab)

        # Selecionar Estado.
        print_step(f"Selecionando estado: {sigla_estado}")
        dropdown_estados = driver.find_element(By.XPATH,"//select[@id='fPP:Decoration:estadoComboOAB']")
        opcoes_estados = Select(dropdown_estados)
        opcoes_estados.select_by_visible_text(sigla_estado)

        # Clicar em Pesquisar.
        print_step("Clicando em Pesquisar...")
        botao_pesquisar = driver.find_element(By.XPATH,"//input[@id='fPP:searchProcessos']")
        botao_pesquisar.click()
        sleep(5)

        # Entrar em cada um dos processos.
        processos = driver.find_elements(By.XPATH,"//b[@class='btn-block']")
        for processo in processos:
            processo.click()
            sleep(5)
            janelas = driver.window_handles
            driver.switch_to.window(janelas[-1])
            driver.set_window_size(1920,1080)
            # Extrair o nº do processo.
            numero_processo = driver.find_elements(By.XPATH,"//div[@class='col-sm-12 ']")
            numero_processo = numero_processo[0]
            numero_processo = numero_processo.text
            # Extrair data de distribuição.
            data_distribuicao = driver.find_elements(By.XPATH,"//div[@class='value col-sm-12 ']")
            data_distribuicao = data_distribuicao [1]
            data_distribuicao = data_distribuicao.text
            # Extrair Classe Judicial.
            classe_judicial = driver.find_elements(By.XPATH,"//div[@class='value col-sm-12 ']")
            classe_judicial = classe_judicial [2]
            classe_judicial = classe_judicial.text
            # Extrair Assunto.
            assunto = driver.find_elements(By.XPATH,"//div[@class='value col-sm-12 ']")
            assunto = assunto [3]
            assunto = assunto.text
            # Extrair Jurisdição.
            jurisdicao = driver.find_elements(By.XPATH,"//div[@class='value col-sm-12 ']")
            jurisdicao = jurisdicao [4]
            jurisdicao = jurisdicao.text
            # Extrair Órgão Julgador.
            orgao_julgador = driver.find_elements(By.XPATH,"//div[@class='value col-sm-12 ']")
            orgao_julgador = orgao_julgador [6]
            orgao_julgador = orgao_julgador.text
            # Extrair e guardar todas as últimas movimentações.
            movimentacoes = driver.find_elements(By.XPATH,"//div[@id='j_id135:processoEventoPanel_body']//tr[contains(@class, 'rich-table-row')]//td//div//div//span")
            lista_movimentacoes = []
            for movimentacao in movimentacoes:
                lista_movimentacoes.append(movimentacao.text)

            print_step(f"Guardando dados do processo {numero_processo} no excel...")
            workbook = openpyxl.load_workbook(caminho_do_excel_raiz)

            try:
             # código para adicionar os dados em página existente
             # acessar página do processo
                pagina_processo = workbook[numero_processo]
            # criar nome das colunas
                pagina_processo['A1'].value = "Número Processo"
                pagina_processo['B1'].value = "Data Distribuição"
                pagina_processo['C1'].value = "Classe Judicial"
                pagina_processo['D1'].value = "Assunto"
                pagina_processo['E1'].value = "Jurisdição"
                pagina_processo['F1'].value = "Órgão Julgador"
                pagina_processo['G1'].value = "Movimentações"
                # adicionar número do processo
                pagina_processo['A2'].value = numero_processo
                # adicionar data de distribuição
                pagina_processo['B2'].value = data_distribuicao
                # adicionar classe judicial
                pagina_processo['C2'].value = classe_judicial
                # adicionar Assunto
                pagina_processo['D2'].value = assunto
                # adicionar jurisdição
                pagina_processo['E2'].value = jurisdicao
                # adicionar orgão julgador
                pagina_processo['F2'].value = orgao_julgador
                # adicionar movimentações
                for index, linha in enumerate(pagina_processo.iter_rows(min_row=2, max_row=len(lista_movimentacoes), min_col=7, max_col=7)):
                    for celula in linha:
                        celula.value = lista_movimentacoes[index]
                workbook.save(caminho_do_excel_raiz)
                print_step(f"Dados salvos com sucesso em {caminho_do_excel_raiz}")
                driver.close()
                sleep(5)
                driver.switch_to.window(driver.window_handles[0])

            except Exception as error:
                # código para criar uma página do zero e adicionar os dados
                workbook.create_sheet(numero_processo)
                # acessar página do processo
                pagina_processo = workbook[numero_processo]
                # criar nome das colunas
                pagina_processo['A1'].value = "Número Processo"
                pagina_processo['B1'].value = "Data Distribuição"
                pagina_processo['C1'].value = "Classe Judicial"
                pagina_processo['D1'].value = "Assunto"
                pagina_processo['E1'].value = "Jurisdição"
                pagina_processo['F1'].value = "Órgão Julgador"
                pagina_processo['G1'].value = "Movimentações"
                # adicionar número do processo
                pagina_processo['A2'].value = numero_processo
                # adicionar data de distribuição
                pagina_processo['B2'].value = data_distribuicao
                # adicionar classe judicial
                pagina_processo['C2'].value = classe_judicial
                # adicionar Assunto
                pagina_processo['D2'].value = assunto
                # adicionar jurisdição
                pagina_processo['E2'].value = jurisdicao
                # adicionar orgão julgador
                pagina_processo['F2'].value = orgao_julgador
                # adicionar movimentações
                for index, linha in enumerate(pagina_processo.iter_rows(min_row=2, max_row=len(lista_movimentacoes), min_col=7, max_col=7)):
                    for celula in linha:
                        celula.value = lista_movimentacoes[index]
                workbook.save(caminho_do_excel_raiz)
                print_step(f"Dados salvos com sucesso em {caminho_do_excel_raiz}")
                driver.close()
                sleep(5)
                driver.switch_to.window(driver.window_handles[0])

        print_step(f"Consultas finalizadas")
    except Exception as e:
        print(f"Erro: {e}")

def abrir_linkedin():
    url_linkedin = "https://www.linkedin.com/in/gabrielnikolax/"
    
    edge_linkedin_options = Options()
    #edge_options.add_argument('--headless')
    edge_linkedin_options.add_experimental_option("detach", True)
    edge_linkedin_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    edge_linkedin_options.add_experimental_option("useAutomationExtension", False)
    edge_linkedin_options.add_argument("--disable-features=IsolateOrigins,site-per-process")

    driver = webdriver.Edge(options=edge_linkedin_options)
    driver.get(url_linkedin)

#Interface Gráfica
customtkinter.set_appearance_mode("light")

# Janela principal
app = customtkinter.CTk(
    fg_color="black"
    )
app.geometry("700x500")
app.resizable(False, False)
app.title("Automatizador de Consulta")

#Frame(Onde ficará o texto)
frame = customtkinter.CTkFrame(
    master= app, 
    width=320, 
    height=360)
frame.place(
    relx=0.5, 
    rely=0.5, 
    anchor=tk.CENTER)

# Título
titulo = customtkinter.CTkLabel(
    master=frame, 
    text="Automatizador de Consulta \nProcessos PJE(TR1)", 
    font=("Century Gothic",20, "bold")
    )
titulo.place(x=22, y=45)

# Número OAB
numero_oab_label = customtkinter.CTkLabel(
    master=frame, 
    text="Número da OAB:", 
    font=("Century Gothic", 15)
    )
numero_oab_label.place(x=20, y=110)
numero_oab_entry = customtkinter.CTkEntry(
    master=frame, 
    placeholder_text="* Preencha apenas números, exemplos: 123456",
    width=280, 
    height=35, 
    placeholder_text_color="red", 
    font=("Century Gothic", 10)
    )
numero_oab_entry.place(x=20, y=135)

# Sigla Estado
sigla_estado_label = customtkinter.CTkLabel(
    master=frame, 
    text="Sigla do Estado:", 
    font=("Century Gothic", 15)
    )
sigla_estado_label.place(x=20, y=185)
sigla_estado_entry = customtkinter.CTkEntry(
    master=frame, 
    placeholder_text="* Preencha apenas as siglas, exemplos: SP, PA, RJ",
    width=280, 
    height=35, 
    placeholder_text_color="red", 
    font=("Century Gothic", 10))
sigla_estado_entry.place(x=20, y=210)

# Botão para executar o scraping
botao_confirmar = customtkinter.CTkButton(
    master=frame, 
    text="Confirmar", 
    width=200, 
    font=("Century Gothic", 18, "bold"), 
    command=scrape_website, 
    corner_radius=200,
    fg_color="#BB1818",
    hover_color="#FF2424"
    )
botao_confirmar.place(x=60, y=290)

# Imagem Botão para Linkedin
img_linkedin = Image.open("img/linkedin.png")
img_linkedin = img_linkedin.resize((32,32))
icone_linkedin = ImageTk.PhotoImage(img_linkedin)

# Botão para o Linkedin
botao_linkedin = customtkinter.CTkButton(
    master= app, 
    command=abrir_linkedin,
    image=icone_linkedin,
    compound=tk.LEFT, 
    text="Linkedin", 
    font=("Century Gothic", 13, "bold"),
    corner_radius=200,
    fg_color="#BB1818",
    hover_color="#FF2424"
    )
botao_linkedin.place(x=550, y=450)

# Loop principal da interface gráfica
app.mainloop()






